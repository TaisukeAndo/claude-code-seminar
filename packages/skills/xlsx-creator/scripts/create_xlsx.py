#!/usr/bin/env python3
"""xlsx-creator: workbook.json → Excel (.xlsx)"""

import json
import sys
import re
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import DataBarRule, Rule
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("openpyxl が必要です: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

SKILL_ROOT = Path(__file__).resolve().parent.parent
ASSETS_DIR = SKILL_ROOT / "assets"

DEFAULT_STYLE = {
    "primary_color": "1E3A5F",
    "secondary_color": "4A90D9",
    "accent_color": "E8700A",
    "header_font": "游ゴシック",
    "body_font": "游ゴシック",
    "header_size": 10,
    "body_size": 10,
}

NUM_FORMATS = {
    "date": "YYYY/MM/DD",
    "date_short": "MM/DD",
    "number": "#,##0",
    "currency": '\\"¥\\"#,##0',
    "percent": "0%",
    "decimal": "#,##0.00",
    "text": "@",
}


# ── Utility helpers ──────────────────────────────────────────────────────────

def to_argb(hex_color: str) -> str:
    h = hex_color.lstrip("#")
    return ("FF" + h).upper() if len(h) == 6 else h.upper()


def mk_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=to_argb(color))


def mk_font(name: str, size=10, bold=False, color="000000", italic=False) -> Font:
    return Font(name=name, size=size, bold=bold, color=to_argb(color), italic=italic)


def mk_side(style="thin", color="CCCCCC") -> Side:
    return Side(style=style, color=to_argb(color))


def mk_border(color="CCCCCC", style="thin") -> Border:
    s = mk_side(style, color)
    return Border(left=s, right=s, top=s, bottom=s)


def mk_bottom_border(color: str, style="medium") -> Border:
    return Border(bottom=Side(style=style, color=to_argb(color)))


def resolve_asset(filename: str):
    if not filename:
        return None
    p = Path(filename)
    if p.is_absolute() and p.exists():
        return p
    c = ASSETS_DIR / filename
    if c.exists():
        return c
    for f in ASSETS_DIR.rglob(Path(filename).name):
        return f
    return None


def build_style(spec_style: dict) -> dict:
    s = {**DEFAULT_STYLE}
    if spec_style:
        s.update({k: v for k, v in spec_style.items() if v})
    return s


# ── Sheet builders ───────────────────────────────────────────────────────────

def add_title_row(ws, title_cfg: dict, style: dict, num_cols: int) -> int:
    """Styled title + subtitle rows. Returns number of rows consumed."""
    if not title_cfg:
        return 0
    rows_used = 0

    text = title_cfg.get("text", "")
    if text:
        ws.row_dimensions[1].height = title_cfg.get("height", 42)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        c = ws.cell(row=1, column=1, value=text)
        c.fill = mk_fill(style["primary_color"])
        c.font = mk_font(style["header_font"], size=16, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        rows_used += 1

    subtitle = title_cfg.get("subtitle", "")
    if subtitle:
        r = rows_used + 1
        ws.row_dimensions[r].height = 20
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        c = ws.cell(row=r, column=1, value=subtitle)
        c.fill = mk_fill(style["secondary_color"])
        c.font = mk_font(style["header_font"], size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        rows_used += 1

    return rows_used


def add_info_rows(ws, info_rows: list, style: dict, start_row: int, num_cols: int) -> int:
    """Project info rows (label/value pairs). Returns rows consumed."""
    for i, info in enumerate(info_rows):
        r = start_row + i
        ws.row_dimensions[r].height = 18
        label = info.get("label", "")
        value = info.get("value", "")

        c_label = ws.cell(row=r, column=1, value=label)
        c_label.fill = mk_fill("EEF2F7")
        c_label.font = mk_font(style["body_font"], bold=True, size=style["body_size"])
        c_label.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Merge label cell
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

        c_val = ws.cell(row=r, column=3, value=value)
        c_val.font = mk_font(style["body_font"], size=style["body_size"])
        c_val.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        if num_cols > 3:
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=num_cols)

    return len(info_rows)


def add_header_row(ws, columns: list, style: dict, row: int):
    ws.row_dimensions[row].height = 24
    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=row, column=ci, value=col.get("header", ""))
        c.fill = mk_fill(style["primary_color"])
        c.font = mk_font(style["header_font"], size=style["header_size"], bold=True, color="FFFFFF")
        c.alignment = Alignment(
            horizontal=col.get("header_align", "center"),
            vertical="center",
            wrap_text=False,
        )
        c.border = Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right=Side(style="thin", color=to_argb(style["primary_color"])),
        )


def add_data_rows(ws, columns: list, rows: list, style: dict, start_row: int):
    alt = mk_fill("F5F8FC")
    normal = mk_fill("FFFFFF")
    bottom_side = Side(style="hair", color="DDDDDD")
    right_side = Side(style="hair", color="DDDDDD")
    cell_border = Border(bottom=bottom_side, right=right_side)

    col_map = {col.get("key"): i + 1 for i, col in enumerate(columns)}

    for ri, row_data in enumerate(rows):
        r = start_row + ri
        ws.row_dimensions[r].height = 20
        row_fill = alt if ri % 2 == 1 else normal

        for ci, col in enumerate(columns, 1):
            key = col.get("key", "")
            formula_tmpl = col.get("formula")

            if formula_tmpl:
                # Replace {row} → actual row number
                val = formula_tmpl.replace("{row}", str(r))
                # Replace {col:key} → column letter
                def repl(m):
                    k = m.group(1)
                    idx = col_map.get(k)
                    return get_column_letter(idx) if idx else m.group(0)
                val = re.sub(r"\{col:(\w+)\}", repl, val)
            else:
                val = row_data.get(key, "")

            c = ws.cell(row=r, column=ci, value=val)
            c.fill = row_fill
            c.font = mk_font(style["body_font"], size=style["body_size"])
            c.alignment = Alignment(
                horizontal=col.get("align", "left"),
                vertical="center",
                wrap_text=col.get("wrap", False),
            )
            c.border = cell_border

            fmt = col.get("format")
            if fmt and fmt in NUM_FORMATS:
                c.number_format = NUM_FORMATS[fmt]

    return len(rows)


def add_summary_row(ws, columns: list, style: dict, data_start: int, num_rows: int):
    r = data_start + num_rows
    ws.row_dimensions[r].height = 22
    top_side = Side(style="medium", color=to_argb(style["primary_color"]))
    bot_side = Side(style="medium", color=to_argb(style["primary_color"]))
    summ_border = Border(top=top_side, bottom=bot_side)

    for ci, col in enumerate(columns, 1):
        col_letter = get_column_letter(ci)
        summary = col.get("summary")

        if ci == 1:
            val = "合計"
        elif summary == "sum":
            val = f"=SUM({col_letter}{data_start}:{col_letter}{r - 1})"
        elif summary == "average":
            val = f"=AVERAGE({col_letter}{data_start}:{col_letter}{r - 1})"
        elif summary == "count":
            val = f"=COUNTA({col_letter}{data_start}:{col_letter}{r - 1})"
        else:
            val = ""

        c = ws.cell(row=r, column=ci, value=val)
        c.fill = mk_fill(style["secondary_color"])
        c.font = mk_font(style["header_font"], size=style["body_size"], bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal=col.get("align", "right"), vertical="center")
        c.border = summ_border

        fmt = col.get("format")
        if fmt and fmt in NUM_FORMATS:
            c.number_format = NUM_FORMATS[fmt]


def apply_dropdowns(ws, columns: list, data_start: int, num_rows: int):
    end = data_start + num_rows - 1
    for ci, col in enumerate(columns, 1):
        items = col.get("dropdown")
        if not items:
            continue
        formula = '"' + ",".join(items) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.showDropDown = False
        dv.errorTitle = "入力エラー"
        dv.error = "リストから選択してください"
        cl = get_column_letter(ci)
        dv.sqref = f"{cl}{data_start}:{cl}{end}"
        ws.add_data_validation(dv)


def apply_conditional_formatting(ws, columns: list, style: dict, data_start: int, num_rows: int):
    end = data_start + num_rows - 1
    for ci, col in enumerate(columns, 1):
        rules = col.get("conditional_formatting", [])
        if not rules:
            continue
        cl = get_column_letter(ci)
        cell_range = f"{cl}{data_start}:{cl}{end}"

        for rule_def in rules:
            rtype = rule_def.get("type", "contains")

            if rtype == "data_bar":
                color = rule_def.get("color", style["secondary_color"]).lstrip("#")
                ws.conditional_formatting.add(
                    cell_range,
                    DataBarRule(start_type="min", end_type="max", color=color),
                )
            elif rtype == "contains":
                value = rule_def.get("value", "")
                fill_color = rule_def.get("fill", "FFFFFF")
                font_color = rule_def.get("font_color")
                dxf_font = mk_font(style["body_font"], color=font_color) if font_color else None
                dxf = DifferentialStyle(fill=mk_fill(fill_color), font=dxf_font)
                r = Rule(
                    type="containsText",
                    operator="containsText",
                    text=value,
                    dxf=dxf,
                )
                r.formula = [f'NOT(ISERROR(SEARCH("{value}",{cl}{data_start})))']
                ws.conditional_formatting.add(cell_range, r)


def set_col_widths(ws, columns: list, rows: list):
    for ci, col in enumerate(columns, 1):
        width = col.get("width")
        if width:
            ws.column_dimensions[get_column_letter(ci)].width = width
        else:
            max_len = len(str(col.get("header", "")))
            key = col.get("key", "")
            for row in rows:
                v = str(row.get(key, ""))
                max_len = max(max_len, len(v))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)


def add_logo(ws, logo_path, cell_ref: str = "A1"):
    try:
        img = XLImage(str(logo_path))
        img.height = 38
        img.width = 120
        ws.add_image(img, cell_ref)
        return True
    except Exception as e:
        print(f"Warning: ロゴ挿入失敗: {e}", file=sys.stderr)
        return False


# ── Free-form cell renderer (帳票向け) ──────────────────────────────────────

def add_free_cells(ws, cells: list, style: dict):
    for spec in cells:
        ref = spec.get("cell")
        if not ref:
            continue

        merge = spec.get("merge")
        if merge:
            ws.merge_cells(merge)

        c = ws[ref]
        formula = spec.get("formula")
        c.value = formula if formula else spec.get("value", "")

        cs = spec.get("style", {})
        if cs.get("fill"):
            c.fill = mk_fill(cs["fill"])
        c.font = mk_font(
            style["body_font"],
            size=cs.get("font_size", style["body_size"]),
            bold=cs.get("bold", False),
            color=cs.get("color", "000000"),
            italic=cs.get("italic", False),
        )
        c.alignment = Alignment(
            horizontal=cs.get("align", "left"),
            vertical=cs.get("valign", "center"),
            wrap_text=cs.get("wrap", False),
        )
        fmt = cs.get("format")
        if fmt:
            c.number_format = NUM_FORMATS.get(fmt, fmt)
        if cs.get("border"):
            c.border = mk_border(cs.get("border_color", "999999"))
        if cs.get("border_bottom"):
            c.border = mk_bottom_border(cs.get("border_color", "999999"))


# ── Main sheet processor ─────────────────────────────────────────────────────

def process_sheet(wb: Workbook, spec: dict, style: dict, logo_path):
    name = spec.get("name", "Sheet")
    ws = wb.create_sheet(title=name)

    tab_color = spec.get("tab_color", style["primary_color"])
    ws.sheet_properties.tabColor = to_argb(tab_color)
    ws.sheet_view.zoomScale = spec.get("zoom", 100)
    if not spec.get("show_gridlines", True):
        ws.sheet_view.showGridLines = False

    orient = spec.get("orientation", "portrait")
    ws.page_setup.orientation = orient
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # ── Form mode (cells array) ──
    free_cells = spec.get("cells")
    if free_cells is not None:
        for col_letter, width in spec.get("column_widths", {}).items():
            ws.column_dimensions[col_letter].width = width
        for row_str, height in spec.get("row_heights", {}).items():
            ws.row_dimensions[int(row_str)].height = height
        if logo_path and spec.get("logo_cell"):
            add_logo(ws, logo_path, spec["logo_cell"])
        for merge_range in spec.get("merged_cells", []):
            try:
                ws.merge_cells(merge_range)
            except Exception:
                pass
        add_free_cells(ws, free_cells, style)
        return ws

    # ── Table mode (columns + rows) ──
    columns = spec.get("columns", [])
    rows = spec.get("rows", [])
    if not columns:
        return ws

    num_cols = len(columns)
    cur = 1

    # Logo (top-right corner)
    if logo_path and spec.get("logo", True):
        logo_col = max(1, num_cols - 1)
        ws.row_dimensions[cur].height = 42
        add_logo(ws, logo_path, f"{get_column_letter(logo_col)}{cur}")

    # Title row
    title_cfg = spec.get("title_row")
    if title_cfg:
        rows_added = add_title_row(ws, title_cfg, style, num_cols)
        cur += rows_added
    elif logo_path and spec.get("logo", True):
        cur += 1

    # Info rows
    info_rows = spec.get("info_rows", [])
    if info_rows:
        cur += add_info_rows(ws, info_rows, style, cur, num_cols)
        cur += 1  # blank separator

    # Column headers
    header_row = cur
    add_header_row(ws, columns, style, header_row)
    cur += 1

    # Freeze panes
    freeze = spec.get("freeze_panes")
    ws.freeze_panes = freeze if freeze else f"A{cur}"

    # Data rows
    data_start = cur
    data_rows = list(rows)
    empty_rows = spec.get("empty_rows", 0)
    if empty_rows > 0:
        data_rows += [{} for _ in range(empty_rows)]

    if data_rows:
        add_data_rows(ws, columns, data_rows, style, data_start)
        cur += len(data_rows)

    num_data = cur - data_start

    # Summary row
    if any(col.get("summary") for col in columns) and num_data > 0:
        add_summary_row(ws, columns, style, data_start, num_data)

    # Column widths
    set_col_widths(ws, columns, rows)

    # Data validation
    if num_data > 0:
        apply_dropdowns(ws, columns, data_start, num_data)
        apply_conditional_formatting(ws, columns, style, data_start, num_data)

    return ws


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Usage: create_xlsx.py <workbook.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    spec_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)

    meta = spec.get("metadata", {})
    style = build_style(spec.get("style"))
    logo_path = resolve_asset(meta.get("logo"))

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.properties.title = meta.get("title", "")
    wb.properties.creator = meta.get("author", "")
    wb.properties.company = meta.get("company", "")
    wb.properties.description = meta.get("description", "")

    for sheet_spec in spec.get("sheets", []):
        process_sheet(wb, sheet_spec, style, logo_path)

    if wb.sheetnames:
        wb.active = wb[wb.sheetnames[0]]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    print(f"✅ Excel を生成しました: {output_path}")
    for name in wb.sheetnames:
        print(f"   📋 {name}")


if __name__ == "__main__":
    main()
