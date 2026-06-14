#!/usr/bin/env python3
"""invoice-creator: document.json + .env + clients/*.json → Excel (.xlsx)"""

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("openpyxl が必要です: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from common import (
    load_env, get_config, load_client, load_document, calc_items,
    format_date_jp, get_doc_label, get_amount_label, ASSETS_DIR, today_str,
)

PRIMARY = "1E3A5F"
ACCENT = "4A90D9"
LIGHT_BG = "F0F4F8"
WHITE = "FFFFFF"
GRAY = "666666"


def argb(h: str) -> str:
    h = h.lstrip("#")
    return ("FF" + h).upper()


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb(color))


def font(name="游ゴシック", size=9, bold=False, color="000000", italic=False) -> Font:
    return Font(name=name, size=size, bold=bold, color=argb(color), italic=italic)


def align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def border_all(color="CCCCCC") -> Border:
    s = Side(style="thin", color=argb(color))
    return Border(left=s, right=s, top=s, bottom=s)


def border_bottom(color="AAAAAA", style="medium") -> Border:
    return Border(bottom=Side(style=style, color=argb(color)))


def set_cell(ws, row, col, value="", bold=False, size=9, color="000000", bg=None,
             h="left", v="center", bdr=None, num_fmt=None, italic=False, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(size=size, bold=bold, color=color, italic=italic)
    c.alignment = align(h, v, wrap)
    if bg:
        c.fill = fill(bg)
    if bdr == "all":
        c.border = border_all()
    elif bdr == "bottom":
        c.border = border_bottom()
    elif bdr == "bottom_light":
        c.border = border_bottom("DDDDDD", "thin")
    if num_fmt:
        c.number_format = num_fmt
    return c


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def generate_invoice_sheet(ws, doc: dict, config: dict, client: dict):
    """見積書・請求書共通レイアウト"""
    doc_type = doc.get("document_type", "invoice")
    company = config["company"]
    bank = config["bank"]
    bank2 = config["bank2"]
    items, subtotal, tax_total, total = calc_items(doc.get("items", []))

    # 列幅設定: A=マージン B=No C=品目 D=品目続 E=数量+単位 F=単価 G=税率 H=金額 I=マージン
    col_w = [1.5, 5, 22, 10, 9, 14, 7, 14, 1.5]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 8  # 上余白

    row = 2
    # ── タイトル行 ─────────────────────────────────────────
    ws.row_dimensions[row].height = 36
    set_cell(ws, row, 2, get_doc_label(doc_type), bold=True, size=20, color=PRIMARY, h="center", v="center")
    merge(ws, row, 2, row, 8)
    row += 1

    ws.row_dimensions[row].height = 6
    row += 1

    # ── 取引先名 + 発行日 ──────────────────────────────────
    ws.row_dimensions[row].height = 22
    client_name = client.get("name", "")
    honorific = client.get("honorific", "御中")
    set_cell(ws, row, 2, f"{client_name}　{honorific}", bold=True, size=12, color="000000", bdr="bottom")
    merge(ws, row, 2, row, 5)
    set_cell(ws, row, 6, "発 行 日", bold=True, size=8, color=GRAY, h="right")
    set_cell(ws, row, 7, format_date_jp(doc.get("issue_date") or today_str()), size=8)
    merge(ws, row, 7, row, 8)
    row += 1

    ws.row_dimensions[row].height = 6
    row += 1

    # 文書番号 + 支払期日
    ws.row_dimensions[row].height = 18
    if doc_type == "invoice":
        set_cell(ws, row, 2, "下記の通りご請求申し上げます。", size=9, color="444444")
    else:
        set_cell(ws, row, 2, "下記の通りお見積り申し上げます。", size=9, color="444444")
    merge(ws, row, 2, row, 5)
    set_cell(ws, row, 6, "文 書 番 号", bold=True, size=8, color=GRAY, h="right")
    set_cell(ws, row, 7, doc.get("document_number", ""), size=8)
    merge(ws, row, 7, row, 8)
    row += 1

    if doc_type == "invoice" and doc.get("due_date"):
        ws.row_dimensions[row].height = 16
        set_cell(ws, row, 6, "お支払期日", bold=True, size=8, color=GRAY, h="right")
        set_cell(ws, row, 7, format_date_jp(doc["due_date"]), size=8)
        merge(ws, row, 7, row, 8)
        row += 1

    ws.row_dimensions[row].height = 8
    row += 1

    # ── 請求金額ボックス ────────────────────────────────────
    ws.row_dimensions[row].height = 30
    set_cell(ws, row, 2, get_amount_label(doc_type), bold=True, size=10, color=WHITE, bg=PRIMARY, h="center", v="center")
    merge(ws, row, 2, row, 3)
    c = ws.cell(row=row, column=4, value=total)
    c.font = Font(name="游ゴシック", size=18, bold=True, color=argb(PRIMARY))
    c.alignment = align("right", "center")
    c.number_format = '¥#,##0'
    merge(ws, row, 4, row, 6)
    set_cell(ws, row, 7, "（消費税込）", size=8, color=GRAY, h="left", v="center")
    merge(ws, row, 7, row, 8)
    # ボックス枠線
    box_color = PRIMARY
    for col in range(2, 9):
        c2 = ws.cell(row=row, column=col)
        c2.border = Border(
            left=Side(style="medium", color=argb(box_color)) if col == 2 else None,
            right=Side(style="medium", color=argb(box_color)) if col == 8 else None,
            top=Side(style="medium", color=argb(box_color)),
            bottom=Side(style="medium", color=argb(box_color)),
        )
    row += 1

    ws.row_dimensions[row].height = 10
    row += 1

    # ── 発行者情報（右側） ─────────────────────────────────
    sender_row = row
    set_cell(ws, row, 6, company.get("name", ""), bold=True, size=10, color=PRIMARY)
    merge(ws, row, 6, row, 8)
    row += 1
    addr = f"〒{company.get('zip', '')} {company.get('address', '')}"
    ws.row_dimensions[row].height = 18
    set_cell(ws, row, 6, addr, size=8, color="333333", wrap=True)
    merge(ws, row, 6, row, 8)
    row += 1
    set_cell(ws, row, 6, f"TEL: {company.get('tel', '')}", size=8, color="333333")
    merge(ws, row, 6, row, 8)
    row += 1
    if company.get("email"):
        set_cell(ws, row, 6, f"Email: {company.get('email', '')}", size=8, color="333333")
        merge(ws, row, 6, row, 8)
        row += 1
    rep = company.get("representative", "")
    rep_title = company.get("representative_title", "")
    if rep:
        set_cell(ws, row, 6, f"{rep_title}　{rep}", size=8, color="333333")
        merge(ws, row, 6, row, 8)
        row += 1
    reg = company.get("registration_number", "")
    if reg:
        set_cell(ws, row, 6, f"登録番号: {reg}", size=8, color=GRAY, italic=True)
        merge(ws, row, 6, row, 8)
        row += 1

    # 印鑑画像
    seal_path = ASSETS_DIR / company.get("seal_image", "seal.png")
    if seal_path.exists():
        try:
            img = XLImage(str(seal_path))
            img.width = 55
            img.height = 55
            ws.add_image(img, f"H{sender_row}")
        except Exception:
            pass

    ws.row_dimensions[row].height = 10
    row += 1

    # ── 件名 ──────────────────────────────────────────────
    if doc.get("title"):
        ws.row_dimensions[row].height = 20
        set_cell(ws, row, 2, "件　名", bold=True, size=9, color=WHITE, bg=PRIMARY, h="center")
        set_cell(ws, row, 3, doc["title"], size=10, bold=True)
        merge(ws, row, 3, row, 8)
        row += 1

    ws.row_dimensions[row].height = 8
    row += 1

    # ── 明細テーブル ヘッダー ──────────────────────────────
    ws.row_dimensions[row].height = 20
    for col, (label, h_align) in enumerate([
        ("No.", "center"), ("品　目　・　内　容", "center"), (None, None),
        ("数量・単位", "center"), ("単　価", "center"), ("税率", "center"), ("金額（税抜）", "center")
    ], start=2):
        if label is not None:
            set_cell(ws, row, col, label, bold=True, size=8, color=WHITE, bg=PRIMARY, h=h_align, v="center")
    merge(ws, row, 3, row, 4)
    row += 1

    # 明細行
    for i, item in enumerate(items, 1):
        ws.row_dimensions[row].height = 18
        bg = WHITE if i % 2 == 1 else "F8F9FA"
        qty_unit = f"{item.get('quantity', 1)} {item.get('unit', '')}"
        set_cell(ws, row, 2, i, size=9, h="center", bg=bg, bdr="all")
        set_cell(ws, row, 3, item.get("name", ""), size=9, bg=bg, bdr="all")
        merge(ws, row, 3, row, 4)
        ws.cell(row=row, column=4).border = border_all()
        set_cell(ws, row, 5, qty_unit, size=9, h="center", bg=bg, bdr="all")
        set_cell(ws, row, 6, item.get("unit_price", 0), size=9, h="right", bg=bg, bdr="all", num_fmt="¥#,##0")
        set_cell(ws, row, 7, f"{item.get('tax_rate', 10)}%", size=9, h="center", bg=bg, bdr="all")
        set_cell(ws, row, 8, item.get("amount", 0), size=9, h="right", bg=bg, bdr="all", num_fmt="¥#,##0")
        row += 1

    # 空行（書き足し用）3行
    for _ in range(3):
        ws.row_dimensions[row].height = 16
        for col in range(2, 9):
            ws.cell(row=row, column=col).border = border_all("EEEEEE")
        merge(ws, row, 3, row, 4)
        row += 1

    ws.row_dimensions[row].height = 8
    row += 1

    # ── 小計・税・合計 ─────────────────────────────────────
    def total_row(label, value, bold=False, bg=None, color="000000"):
        nonlocal row
        ws.row_dimensions[row].height = 18
        set_cell(ws, row, 6, label, bold=bold, size=9, color=color, bg=bg or LIGHT_BG, h="right")
        merge(ws, row, 6, row, 7)
        ws.cell(row=row, column=7).fill = fill(bg or LIGHT_BG)
        c = ws.cell(row=row, column=8, value=value)
        c.font = Font(name="游ゴシック", size=9, bold=bold, color=argb(color))
        c.alignment = align("right")
        c.number_format = "¥#,##0"
        c.fill = fill(bg or LIGHT_BG)
        c.border = border_all()
        row += 1

    total_row("小計（税抜）", subtotal)
    total_row("消費税（10%）", tax_total)

    # 合計行（強調）
    ws.row_dimensions[row].height = 22
    set_cell(ws, row, 6, "合計（税込）", bold=True, size=10, color=WHITE, bg=PRIMARY, h="right", v="center")
    merge(ws, row, 6, row, 7)
    ws.cell(row=row, column=7).fill = fill(PRIMARY)
    c = ws.cell(row=row, column=8, value=total)
    c.font = Font(name="游ゴシック", size=11, bold=True, color=argb(WHITE))
    c.alignment = align("right", "center")
    c.number_format = "¥#,##0"
    c.fill = fill(PRIMARY)
    row += 1

    ws.row_dimensions[row].height = 10
    row += 1

    # ── 振込先（請求書のみ） ───────────────────────────────
    if doc_type == "invoice" and bank.get("name"):
        ws.row_dimensions[row].height = 20
        set_cell(ws, row, 2, "お振込先", bold=True, size=9, color=WHITE, bg=ACCENT, h="center", v="center")
        merge(ws, row, 2, row, 3)
        bank_str = f"{bank['name']}　{bank['branch']}　{bank['type']}預金　{bank['number']}　口座名義: {bank['holder']}"
        set_cell(ws, row, 4, bank_str, size=9, color="222222", bg=LIGHT_BG, v="center")
        merge(ws, row, 4, row, 8)
        row += 1
        if bank2.get("name"):
            ws.row_dimensions[row].height = 18
            set_cell(ws, row, 2, "第2口座", bold=True, size=9, color=WHITE, bg=ACCENT, h="center", v="center")
            merge(ws, row, 2, row, 3)
            bank2_str = f"{bank2['name']}　{bank2['branch']}　{bank2['type']}預金　{bank2['number']}　口座名義: {bank2['holder']}"
            set_cell(ws, row, 4, bank2_str, size=9, color="222222", bg=LIGHT_BG, v="center")
            merge(ws, row, 4, row, 8)
            row += 1
        ws.row_dimensions[row].height = 8
        row += 1

    # ── 備考 ──────────────────────────────────────────────
    if doc.get("notes"):
        ws.row_dimensions[row].height = 20
        set_cell(ws, row, 2, "備　考", bold=True, size=9, color=WHITE, bg=GRAY, h="center", v="center")
        merge(ws, row, 2, row, 3)
        set_cell(ws, row, 4, doc["notes"], size=9, color="333333", bg=LIGHT_BG, v="center", wrap=True)
        merge(ws, row, 4, row, 8)
        row += 1


def generate_receipt_sheet(ws, doc: dict, config: dict, client: dict):
    """領収書レイアウト"""
    company = config["company"]
    total_amount = doc.get("total_amount", 0)
    tax_amount = doc.get("tax_amount", 0)
    subtotal_amount = doc.get("subtotal_amount", total_amount - tax_amount)

    col_w = [2, 14, 8, 8, 14, 8, 14, 8, 2]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    ws.row_dimensions[row].height = 36
    set_cell(ws, row, 2, "領　収　書", bold=True, size=20, color=PRIMARY, h="center", v="center")
    merge(ws, row, 2, row, 8)
    row += 1

    ws.row_dimensions[row].height = 8
    row += 1

    # 発行日・文書番号
    ws.row_dimensions[row].height = 18
    set_cell(ws, row, 6, "発 行 日", bold=True, size=8, color=GRAY, h="right")
    set_cell(ws, row, 7, format_date_jp(doc.get("issue_date") or today_str()), size=9)
    merge(ws, row, 7, row, 8)
    row += 1
    ws.row_dimensions[row].height = 16
    set_cell(ws, row, 6, "文書番号", bold=True, size=8, color=GRAY, h="right")
    set_cell(ws, row, 7, doc.get("document_number", ""), size=9)
    merge(ws, row, 7, row, 8)
    row += 1

    ws.row_dimensions[row].height = 10
    row += 1

    # 支払者名
    ws.row_dimensions[row].height = 24
    client_name = client.get("name", "")
    honorific = client.get("honorific", "様")
    set_cell(ws, row, 2, f"{client_name}　{honorific}", bold=True, size=13, color="000000", bdr="bottom")
    merge(ws, row, 2, row, 5)
    row += 1

    ws.row_dimensions[row].height = 8
    row += 1

    # 金額ボックス
    ws.row_dimensions[row].height = 32
    set_cell(ws, row, 2, "お受取金額", bold=True, size=10, color=WHITE, bg=PRIMARY, h="center", v="center")
    merge(ws, row, 2, row, 3)
    c = ws.cell(row=row, column=4, value=total_amount)
    c.font = Font(name="游ゴシック", size=20, bold=True, color=argb(PRIMARY))
    c.alignment = align("right", "center")
    c.number_format = "¥#,##0"
    merge(ws, row, 4, row, 6)
    set_cell(ws, row, 7, "（消費税込）", size=8, color=GRAY, h="left", v="center")
    merge(ws, row, 7, row, 8)
    for col in range(2, 9):
        ws.cell(row=row, column=col).border = Border(
            left=Side(style="medium", color=argb(PRIMARY)) if col == 2 else None,
            right=Side(style="medium", color=argb(PRIMARY)) if col == 8 else None,
            top=Side(style="medium", color=argb(PRIMARY)),
            bottom=Side(style="medium", color=argb(PRIMARY)),
        )
    row += 1

    ws.row_dimensions[row].height = 8
    row += 1

    # 発行者情報（右）
    sender_row = row
    set_cell(ws, row, 6, company.get("name", ""), bold=True, size=10, color=PRIMARY)
    merge(ws, row, 6, row, 8)
    row += 1
    addr = f"〒{company.get('zip', '')} {company.get('address', '')}"
    ws.row_dimensions[row].height = 18
    set_cell(ws, row, 6, addr, size=8, color="333333", wrap=True)
    merge(ws, row, 6, row, 8)
    row += 1
    set_cell(ws, row, 6, f"TEL: {company.get('tel', '')}", size=8, color="333333")
    merge(ws, row, 6, row, 8)
    row += 1
    rep = company.get("representative", "")
    if rep:
        set_cell(ws, row, 6, f"{company.get('representative_title', '')}　{rep}", size=8, color="333333")
        merge(ws, row, 6, row, 8)
        row += 1
    reg = company.get("registration_number", "")
    if reg:
        set_cell(ws, row, 6, f"登録番号: {reg}", size=8, color=GRAY, italic=True)
        merge(ws, row, 6, row, 8)
        row += 1

    seal_path = ASSETS_DIR / company.get("seal_image", "seal.png")
    if seal_path.exists():
        try:
            img = XLImage(str(seal_path))
            img.width = 55
            img.height = 55
            ws.add_image(img, f"H{sender_row}")
        except Exception:
            pass

    ws.row_dimensions[row].height = 10
    row += 1

    # 但し書き
    if doc.get("title"):
        ws.row_dimensions[row].height = 20
        set_cell(ws, row, 2, "但し書き", bold=True, size=9, color=WHITE, bg=PRIMARY, h="center")
        set_cell(ws, row, 3, doc["title"] + "　として", size=10)
        merge(ws, row, 3, row, 8)
        row += 1

    ws.row_dimensions[row].height = 8
    row += 1

    # 内訳
    ws.row_dimensions[row].height = 20
    set_cell(ws, row, 2, "内　訳", bold=True, size=9, color=WHITE, bg=GRAY, h="center", v="center")
    merge(ws, row, 2, row, 3)
    set_cell(ws, row, 4, "税抜金額", size=8, color=GRAY, h="right", bg=LIGHT_BG)
    merge(ws, row, 4, row, 5)
    c = ws.cell(row=row, column=6, value=subtotal_amount)
    c.font = font(size=9)
    c.alignment = align("right")
    c.number_format = "¥#,##0"
    c.fill = fill(LIGHT_BG)
    set_cell(ws, row, 7, "消費税", size=8, color=GRAY, h="right", bg=LIGHT_BG)
    c = ws.cell(row=row, column=8, value=tax_amount)
    c.font = font(size=9)
    c.alignment = align("right")
    c.number_format = "¥#,##0"
    c.fill = fill(LIGHT_BG)
    row += 1

    # 収入印紙注記
    if total_amount >= 50000:
        ws.row_dimensions[row].height = 10
        row += 1
        ws.row_dimensions[row].height = 16
        set_cell(ws, row, 2, "※ この領収書は50,000円以上のため収入印紙が必要な場合があります。",
                 size=8, color="CC0000", italic=True)
        merge(ws, row, 2, row, 8)
        row += 1

    if doc.get("notes"):
        ws.row_dimensions[row].height = 10
        row += 1
        ws.row_dimensions[row].height = 20
        set_cell(ws, row, 2, "備　考", bold=True, size=9, color=WHITE, bg=GRAY, h="center", v="center")
        merge(ws, row, 2, row, 3)
        set_cell(ws, row, 4, doc["notes"], size=9, color="333333", bg=LIGHT_BG, v="center")
        merge(ws, row, 4, row, 8)


def main():
    if len(sys.argv) < 3:
        print("使い方: python3 generate_excel.py <document.json> <output.xlsx>")
        sys.exit(1)

    doc_path = sys.argv[1]
    out_path = sys.argv[2]

    env = load_env()
    config = get_config(env)
    doc = load_document(doc_path)
    client = load_client(doc.get("client_id", ""))

    wb = Workbook()
    ws = wb.active
    doc_type = doc.get("document_type", "invoice")
    ws.title = {"estimate": "見積書", "invoice": "請求書", "receipt": "領収書"}.get(doc_type, "請求書")

    if doc_type == "receipt":
        generate_receipt_sheet(ws, doc, config, client)
    else:
        generate_invoice_sheet(ws, doc, config, client)

    # 印刷設定
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.7
    ws.page_margins.bottom = 0.7

    wb.save(out_path)
    doc_type_label = {"estimate": "見積書", "invoice": "請求書", "receipt": "領収書"}.get(doc_type, "書類")
    print(f"✅ {doc_type_label}（Excel）生成完了: {out_path}")


if __name__ == "__main__":
    main()
